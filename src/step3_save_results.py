"""
Step 3: Save review results to Excel with formatting
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from pathlib import Path
from models import ReviewTable
from logger import logger


def save_to_excel(
    review_table: ReviewTable,
    round_number: int,
    output_dir: str
) -> str:
    """
    Save review results to formatted Excel file

    Args:
        review_table: AI review results
        round_number: Round number (e.g., 6 for "第六回")
        output_dir: Output directory

    Returns:
        str: Path to generated Excel file

    Raises:
        ValueError: If review_table is empty
        Exception: If Excel generation fails
    """
    if not review_table.rows:
        raise ValueError("Review table is empty. No data to save.")

    # Create output directory
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Convert number to Japanese round format
    japanese_numbers = {
        1: "一", 2: "二", 3: "三", 4: "四", 5: "五",
        6: "六", 7: "七", 8: "八", 9: "九", 10: "十"
    }
    round_name = japanese_numbers.get(round_number, str(round_number))
    filename = f"第{round_name}回.xlsx"
    output_file = output_path / filename

    logger.info(f"Creating Excel file: {filename}")
    logger.info(f"Processing {len(review_table.rows)} rows")

    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "評審結果"

        # Define column headers
        headers = [
            "要求No",
            "要求内容 (ペルソナ: 指令)",
            "評価 (〇/△/×)",
            "適合/不適合箇所",
            "適合/不適合理由",
            "修正案 (ゴールデンケースを含む)",
            "対応有無",
            "対応方法／非対応理由"
        ]

        # Write header row
        ws.append(headers)

        # Write data rows
        for row in review_table.rows:
            ws.append([
                row.requirement_no,
                row.requirement_content,
                row.evaluation,
                row.compliance_location,
                row.compliance_reason,
                row.correction_plan,
                row.response_status,
                row.response_method
            ])

        logger.info("Data written to worksheet")

        # === Apply Formatting ===

        # 1. Header row: Light blue background
        header_fill = PatternFill(
            start_color="ADD8E6",
            end_color="ADD8E6",
            fill_type="solid"
        )
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 2. All cells: Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                               min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                # Align data cells
                if cell.row > 1:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=True
                    )

        # 3. AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # 4. Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    # Consider line breaks
                    lines = cell_value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # Set column width (minimum 10, maximum 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 5. Set row height for wrapped text
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = None  # Auto height

        logger.info("Formatting applied")

        # Save workbook
        wb.save(output_file)
        logger.info(f"Excel file saved successfully: {output_file}")

        return str(output_file)

    except Exception as e:
        logger.error(f"Failed to save Excel file: {str(e)}")
        raise


if __name__ == "__main__":
    # Test code
    from models import ReviewRow

    # Create sample data
    sample_rows = [
        ReviewRow(
            requirement_no="REQ-001",
            requirement_content="システムは24時間稼働すること",
            evaluation="〇",
            compliance_location="仕様書 3.2節",
            compliance_reason="要件を満たしている",
            correction_plan="特になし",
            response_status="対応済",
            response_method="既に実装済み"
        ),
        ReviewRow(
            requirement_no="REQ-002",
            requirement_content="応答時間は1秒以内であること",
            evaluation="△",
            compliance_location="パフォーマンステスト結果",
            compliance_reason="平均1.2秒で要件未達",
            correction_plan="データベースクエリの最適化を実施",
            response_status="対応中",
            response_method="次回リリースで対応予定"
        )
    ]

    sample_table = ReviewTable(rows=sample_rows)

    try:
        output_path = save_to_excel(sample_table, 6, "./output/results")
        print(f"Test successful: {output_path}")
    except Exception as e:
        print(f"Test failed: {str(e)}")
